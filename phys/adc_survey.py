#!/usr/bin/env python3
"""Plot area trade-offs from the Murmann ADC performance survey."""

from __future__ import annotations

import argparse
import math
import os
from collections.abc import Callable, Sequence
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

os.environ.setdefault("MPLBACKEND", "Agg")

import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
from openpyxl import load_workbook

from flow.analysis.plots import (
    CURVE_COLORS,
    FULL_HD_FIGSIZE,
    LEGEND_FACE_COLOR,
    PLOT_STYLE,
    SPINE_COLOR,
    TEXT_COLOR,
    _save_figure,
    style_ax,
    style_grid,
    style_legend,
)

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "libs" / "ADC-survey" / "xls" / "ADCsurvey_latest.xlsx"
SURVEY_SHEETS = ("ISSCC", "VLSI")

AREA_MAX_MM2 = 0.25
ENOB_MIN = 6.0
NYQUIST_RATE_MIN_HZ = 0.1e6
CONVERSION_ENERGY_MAX_PJ = 1_000.0
FILTER_TOLERANCE = 1e-9

FRIDA_TARGET_AREA_UM2 = 60.0 * 60.0
FRIDA_TARGET_TECHNOLOGY_NM = 65.0
FRIDA_TARGET_ENOB = 11.0
FRIDA_TARGET_DESIGN_BITS = 12
FRIDA_TARGET_POWER_W = 100e-6
FRIDA_TARGET_SAMPLING_RATE_HZ = 10e6
FRIDA_TARGET_SNDR_DB = 6.02 * FRIDA_TARGET_ENOB + 1.76
FRIDA_TARGET_CONVERSION_ENERGY_PJ = FRIDA_TARGET_POWER_W / FRIDA_TARGET_SAMPLING_RATE_HZ * 1e12
FRIDA_TARGET_WALDEN_FOM_FJ = FRIDA_TARGET_POWER_W / (FRIDA_TARGET_SAMPLING_RATE_HZ * 2**FRIDA_TARGET_ENOB) * 1e15

TECHNOLOGY_COLORS = {
    "≤ 16 nm": CURVE_COLORS[0],
    "20 / 22 nm": CURVE_COLORS[1],
    "28 / 32 nm": CURVE_COLORS[2],
    "40 / 45 nm": CURVE_COLORS[3],
    "55 / 65 nm": CURVE_COLORS[4],
    "≥ 90 nm": CURVE_COLORS[5],
}
ARCHITECTURE_MARKERS = {
    "SAR": "o",
    "Pipeline SAR": "s",
    "Pipeline": "v",
    "CT ΔΣ": "^",
    "DT / incrmntl. ΔΣ": "D",
    "Slope / ramp": "X",
    "VCO / time-based": "P",
    "Other": "h",
}


@dataclass(frozen=True)
class AdcSurveyPoint:
    """One filtered operating point from the ADC survey."""

    conference: str
    year: int
    paper_id: str
    architecture: str
    architecture_family: str
    technology_nm: float
    area_mm2: float
    sndr_plot_db: float
    enob: float
    power_w: float
    sampling_rate_hz: float
    nyquist_rate_hz: float
    conversion_energy_pj: float
    walden_fom_fj: float


def _as_float(value: object) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, str):
        try:
            value = float(value.strip())
        except ValueError:
            return None
    if not isinstance(value, (int, float)):
        return None
    converted = float(value)
    return converted if math.isfinite(converted) else None


def technology_category(technology_nm: float) -> str:
    """Return the requested discrete technology-node color category."""

    if technology_nm <= 16.0:
        return "≤ 16 nm"
    if math.isclose(technology_nm, 20.0) or math.isclose(technology_nm, 22.0):
        return "20 / 22 nm"
    if math.isclose(technology_nm, 28.0) or math.isclose(technology_nm, 32.0):
        return "28 / 32 nm"
    if any(math.isclose(technology_nm, node) for node in (38.0, 40.0, 45.0)):
        return "40 / 45 nm"
    if 55.0 <= technology_nm <= 65.0:
        return "55 / 65 nm"
    if technology_nm >= 90.0:
        return "≥ 90 nm"
    raise ValueError(f"technology node has no configured color category: {technology_nm:g} nm")


def architecture_family(architecture: str) -> str:
    """Collapse the survey's detailed architecture tags into marker families."""

    normalized = architecture.casefold()
    if "slope" in normalized:
        return "Slope / ramp"
    if "sdct" in normalized:
        return "CT ΔΣ"
    if any(tag in normalized for tag in ("sdsc", "sddt", "incremental")):
        return "DT / incrmntl. ΔΣ"
    if "pipe" in normalized and "sar" in normalized:
        return "Pipeline SAR"
    if "sar" in normalized:
        return "SAR"
    if "pipe" in normalized:
        return "Pipeline"
    if any(tag in normalized for tag in ("vco", "time-based", "time based")):
        return "VCO / time-based"
    return "Other"


def load_filtered_points(workbook_path: Path = DEFAULT_WORKBOOK) -> tuple[AdcSurveyPoint, ...]:
    """Load survey points satisfying the agreed area, ENOB, rate, and energy cuts."""

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    points: list[AdcSurveyPoint] = []
    try:
        for conference in SURVEY_SHEETS:
            sheet = workbook[conference]
            headers = tuple(cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1)))
            column = {str(name): index for index, name in enumerate(headers) if name is not None}
            required = {
                "YEAR",
                "ID",
                "ARCHITECTURE",
                "TECHNOLOGY",
                "AREA [mm^2]",
                "SNDR_plot [dB]",
                "P [W]",
                "fs [Hz]",
                "fsnyq [Hz]",
            }
            missing = required - column.keys()
            if missing:
                raise ValueError(f"{conference} sheet is missing columns: {sorted(missing)}")

            for row in sheet.iter_rows(min_row=2, values_only=True):
                area_mm2 = _as_float(row[column["AREA [mm^2]"]])
                sndr_plot_db = _as_float(row[column["SNDR_plot [dB]"]])
                power_w = _as_float(row[column["P [W]"]])
                sampling_rate_hz = _as_float(row[column["fs [Hz]"]])
                nyquist_rate_hz = _as_float(row[column["fsnyq [Hz]"]])
                technology_um = _as_float(row[column["TECHNOLOGY"]])
                year = _as_float(row[column["YEAR"]])
                paper_id = row[column["ID"]]
                architecture = row[column["ARCHITECTURE"]]
                numeric = (area_mm2, sndr_plot_db, power_w, sampling_rate_hz, nyquist_rate_hz, technology_um, year)
                if (
                    any(value is None for value in numeric)
                    or paper_id in (None, "")
                    or not isinstance(architecture, str)
                ):
                    continue
                assert area_mm2 is not None
                assert sndr_plot_db is not None
                assert power_w is not None
                assert sampling_rate_hz is not None
                assert nyquist_rate_hz is not None
                assert technology_um is not None
                assert year is not None
                if min(area_mm2, power_w, sampling_rate_hz, nyquist_rate_hz, technology_um) <= 0.0:
                    continue

                enob = (sndr_plot_db - 1.76) / 6.02
                conversion_energy_pj = power_w / nyquist_rate_hz * 1e12
                selected = (
                    area_mm2 <= AREA_MAX_MM2
                    and enob + FILTER_TOLERANCE >= ENOB_MIN
                    and nyquist_rate_hz >= NYQUIST_RATE_MIN_HZ
                    and conversion_energy_pj <= CONVERSION_ENERGY_MAX_PJ
                )
                if not selected:
                    continue

                points.append(
                    AdcSurveyPoint(
                        conference=conference,
                        year=int(year),
                        paper_id=str(paper_id),
                        architecture=architecture,
                        architecture_family=architecture_family(architecture),
                        technology_nm=technology_um * 1e3,
                        area_mm2=area_mm2,
                        sndr_plot_db=sndr_plot_db,
                        enob=enob,
                        power_w=power_w,
                        sampling_rate_hz=sampling_rate_hz,
                        nyquist_rate_hz=nyquist_rate_hz,
                        conversion_energy_pj=conversion_energy_pj,
                        walden_fom_fj=power_w / (nyquist_rate_hz * 2**enob) * 1e15,
                    )
                )
    finally:
        workbook.close()

    identities = {(point.conference, point.year, point.paper_id) for point in points}
    if len(identities) != len(points):
        raise ValueError("filtered survey contains repeated operating points for one conference paper")
    return tuple(points)


def plot_tradeoff(
    points: Sequence[AdcSurveyPoint],
    *,
    x_metric: Callable[[AdcSurveyPoint], float],
    y_metric: Callable[[AdcSurveyPoint], float],
    xlabel: str,
    ylabel: str,
    title: str,
    output_path: Path,
    target_x: float,
    target_y: float,
    xscale: str = "linear",
    yscale: str = "linear",
    xlim: tuple[float, float] | None = None,
) -> tuple[Path, ...]:
    """Plot two ADC metrics using technology colors and architecture markers."""

    if not points:
        raise ValueError("ADC trade-off plot requires at least one point")
    with plt.rc_context(PLOT_STYLE):
        fig, ax = plt.subplots(figsize=FULL_HD_FIGSIZE)
        for family, marker in ARCHITECTURE_MARKERS.items():
            selected = [point for point in points if point.architecture_family == family]
            if not selected:
                continue
            ax.scatter(
                [x_metric(point) for point in selected],
                [y_metric(point) for point in selected],
                color=[TECHNOLOGY_COLORS[technology_category(point.technology_nm)] for point in selected],
                marker=marker,
                s=58,
                edgecolors=SPINE_COLOR,
                linewidths=0.45,
                zorder=3,
            )

        ax.scatter(
            target_x,
            target_y,
            marker=ARCHITECTURE_MARKERS["SAR"],
            s=180,
            color=TECHNOLOGY_COLORS[technology_category(FRIDA_TARGET_TECHNOLOGY_NM)],
            edgecolors=TEXT_COLOR,
            linewidths=1.2,
            zorder=7,
        )
        ax.annotate(
            "Design\ngoal",
            (target_x, target_y),
            xytext=(0, 10),
            textcoords="offset points",
            horizontalalignment="center",
            verticalalignment="bottom",
            color=TEXT_COLOR,
            zorder=8,
        )

        architecture_handles = [
            Line2D(
                [0],
                [0],
                linestyle="none",
                marker=marker,
                markersize=6.5,
                markerfacecolor=LEGEND_FACE_COLOR,
                markeredgecolor=SPINE_COLOR,
                markeredgewidth=0.7,
                label=family,
            )
            for family, marker in ARCHITECTURE_MARKERS.items()
            if any(point.architecture_family == family for point in points)
        ]
        style_legend(
            ax,
            handles=architecture_handles,
            title="Architecture",
            loc="upper left",
        )
        architecture_legend = ax.get_legend()
        if architecture_legend is not None:
            ax.add_artist(architecture_legend)

        present_technology_categories = {technology_category(point.technology_nm) for point in points}
        technology_handles = [
            Line2D(
                [0],
                [0],
                linestyle="none",
                marker="o",
                markersize=7,
                markerfacecolor=color,
                markeredgecolor=SPINE_COLOR,
                markeredgewidth=0.5,
                label=category,
            )
            for category, color in TECHNOLOGY_COLORS.items()
            if category in present_technology_categories
        ]
        style_legend(
            ax,
            handles=technology_handles,
            title="Process Node",
            loc="lower right",
        )

        ax.set_xscale(xscale)
        ax.set_yscale(yscale)
        if xlim is not None:
            ax.set_xlim(*xlim)
        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel)
        style_ax(ax)
        style_grid(ax)
        fig.suptitle(title)
        return _save_figure(fig, output_path)


def generate_plots(
    workbook_path: Path = DEFAULT_WORKBOOK,
    *,
    output_dir: Path,
) -> tuple[Path, ...]:
    """Generate the requested area and energy trade-off plots."""

    points = load_filtered_points(workbook_path)
    outputs = []
    outputs.extend(
        plot_tradeoff(
            points,
            x_metric=lambda point: point.area_mm2 * 1e6,
            y_metric=lambda point: point.walden_fom_fj,
            xlabel="Reported ADC area (µm²)",
            ylabel="Walden FoM (fJ/conversion-step)",
            title="ADC Walden FoM vs reported area",
            output_path=output_dir / "adc_survey_fomw_vs_area",
            target_x=FRIDA_TARGET_AREA_UM2,
            target_y=FRIDA_TARGET_WALDEN_FOM_FJ,
            xscale="log",
            yscale="log",
        )
    )
    outputs.extend(
        plot_tradeoff(
            points,
            x_metric=lambda point: point.area_mm2 * 1e6,
            y_metric=lambda point: point.conversion_energy_pj,
            xlabel="Reported ADC area (µm²)",
            ylabel="Energy per Nyquist conversion, P / fsnyq (pJ)",
            title="ADC conversion energy vs reported area",
            output_path=output_dir / "adc_survey_energy_vs_area",
            target_x=FRIDA_TARGET_AREA_UM2,
            target_y=FRIDA_TARGET_CONVERSION_ENERGY_PJ,
            xscale="log",
            yscale="log",
        )
    )
    outputs.extend(
        plot_tradeoff(
            points,
            x_metric=lambda point: point.area_mm2 * 1e6,
            y_metric=lambda point: point.enob,
            xlabel="Reported ADC area (µm²)",
            ylabel="Effective number of bits (ENOB)",
            title="ADC effective resolution vs reported area",
            output_path=output_dir / "adc_survey_enob_vs_area",
            target_x=FRIDA_TARGET_AREA_UM2,
            target_y=FRIDA_TARGET_ENOB,
            xscale="log",
        )
    )
    outputs.extend(
        plot_tradeoff(
            points,
            x_metric=lambda point: point.conversion_energy_pj,
            y_metric=lambda point: point.enob,
            xlabel="Energy per Nyquist conversion, P / fsnyq (pJ)",
            ylabel="Effective number of bits (ENOB)",
            title="ADC effective resolution vs conversion energy",
            output_path=output_dir / "adc_survey_enob_vs_energy",
            target_x=FRIDA_TARGET_CONVERSION_ENERGY_PJ,
            target_y=FRIDA_TARGET_ENOB,
            xscale="log",
        )
    )
    outputs.extend(
        plot_tradeoff(
            points,
            x_metric=lambda point: point.nyquist_rate_hz / 1e6,
            y_metric=lambda point: point.enob,
            xlabel="Effective Nyquist conversion rate, fsnyq (MS/s)",
            ylabel="Effective number of bits (ENOB)",
            title="ADC effective resolution vs conversion rate",
            output_path=output_dir / "adc_survey_enob_vs_rate",
            target_x=FRIDA_TARGET_SAMPLING_RATE_HZ / 1e6,
            target_y=FRIDA_TARGET_ENOB,
            xscale="log",
        )
    )
    return tuple(outputs)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output-dir", type=Path)
    args = parser.parse_args()
    output_dir = args.output_dir or ROOT / "build" / "adc_survey" / datetime.now().astimezone().strftime("%Y%m%d_%H%M")
    points = load_filtered_points(args.workbook)
    print(f"Selected {len(points)} ADCs")
    for path in generate_plots(args.workbook, output_dir=output_dir):
        print(path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
