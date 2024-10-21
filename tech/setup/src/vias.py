# -*- coding: utf-8 -*-
"""
Created on Fri Mar 25 11:06:32 2016

@author: Koen
"""
import os
import time
import math
# consider xlwings instead of openpyxl if formulas are to be evaluated
import openpyxl

import logging      # in case you want to add extra logging
import general
import LTBsettings


class ViasError(Exception):
    pass


class Layer:
    def __init__(self, name, leditname, leditpurpose, level):
        self.name = name
        self.LEditname = leditname
        self.LEditpurpose = leditpurpose
        self.level = level
        self.width = 0

    def __str__(self):
        return (self.name + ' [' + str(self.level) + '] (L-Edit: ' +
                self.LEditname + ', ' + str(self.LEditpurpose) + ')')

    def __lt__(self, other):
        return self.level < other.level

    def set_width(self, width):
        self.width = width

    def export_autogen(self):
        batchtext = 'LLayer ' + self.name + ';\n'
        if self.LEditpurpose is None:
            batchtext += (self.name + ' = LLayer_Find(activefile,"' +
                          self.LEditname + '");\n')
        else:
            batchtext += (self.name + ' = LLayer_FindByNames(activefile,"' +
                          self.LEditname + '", "' + self.LEditpurpose +
                          '");\n')
        return batchtext

    def isvia(self):
        return isinstance(self, Via)

    def isconductor(self):
        return isinstance(self, Conductor)


class Via(Layer):
    def __init__(self, name, leditname, leditpurpose, level):
        super().__init__(name, leditname, leditpurpose, level)
        self.space = 0
        self.spacefarm = 0
        self.maxdimsizenofarm = 0
        self.enc = {}

    def set_space(self, space):
        self.space = space

    def set_spacefarm(self, spacefarm):
        self.spacefarm = spacefarm

    def set_maxdimsizenofarm(self, maxdimsizenofarm):
        self.maxdimsizenofarm = maxdimsizenofarm

    def add_enclosure(self, layername, enc, enc_2opp):
        self.enc[layername] = (enc, enc_2opp)


class Conductor(Layer):
    def __init__(self, name, leditname, leditpurpose, level):
        super().__init__(name, leditname, leditpurpose, level)
        self.toplevel = False
        self.minarea = 0
        self.widthrelax = 0
        self.horizontal = True

    def set_widthrelax(self, grid):
        self.widthrelax = (math.floor(self.width / grid) + 1) * grid

    def set_minarea(self, minarea):
        self.minarea = minarea

    def set_direction(self, direction):
        if isinstance(direction, str) and direction[0].upper() == 'H':
            self.horizontal = True
        elif isinstance(direction, str) and direction[0].upper() == 'V':
            self.horizontal = False
        else:
            raise Exception('Unknown direction:' + repr(direction) + ' ' +
                            repr(direction.__class__))


class AllLayers:
    def __init__(self):
        self.content = []

    def add(self, name, leditname, leditpurpose, level):
        if 0.25 < level % 1 < .75:
            self.content.append(Via(name, leditname, leditpurpose, level))
        else:
            self.content.append(Conductor(name, leditname, leditpurpose, level))

    def define_toplevel(self):
        toplevel = -1
        for layer in self.content:
            if isinstance(layer, Conductor):
                toplevel = max(toplevel, layer.level)

        check = 0
        for layer in self.content:
            if isinstance(layer, Conductor):
                if layer.level == toplevel:
                    layer.toplevel = True
                    check += 1
        assert check == 1

    def export_autogen(self):
        batchtext = ''
        for layer in self.content:
            batchtext += layer.export_autogen()
        return batchtext


class Polygon:
    def __init__(self, layer, coords):
        assert isinstance(layer, Layer)
        assert len(coords) > 2
        for xy in coords:
            assert len(xy) == 2
        self.layer = layer
        self.coords = coords

    def __str__(self):
        leditgrid = 1  # units/nm
        logtext = 'P '
        for xy in self.coords:
            logtext += ('(' + str(int(round(xy[0]*leditgrid))) + ',' +
                        str(int(round(xy[1]*leditgrid))) + ') ')
        return self.layer.name + ': ' + logtext + '\n'

    def export_autogen(self):
        leditgrid = 1  # units/nm
        batchtext = ''
        points = 0
        for xy in self.coords:
            batchtext += ('\t\tPolygon [' + str(points) + '] = LPoint_Set ( ' +
                          str(int(round(xy[0]*leditgrid))) + ',' +
                          str(int(round(xy[1]*leditgrid))) + ' );\n')
            points += 1
        batchtext += ('\t\tLPolygon_New( newCell, ' + self.layer.name +
                      ', Polygon, ' + str(points) + ');\n')
        return batchtext

    def __eq__(self, other):
        if not(isinstance(other, Polygon)):
            return False
        if self.layer != other.layer:
            return False
        if len(self.coords) != len(other.coords):
            return False
        for loop in range(len(self.coords)):
            if self.coords == other.coords[loop:] + other.coords[:loop]:
                return True
        return False


class Box:
    def __init__(self, layer, coords):
        assert isinstance(layer, Layer)
        assert len(coords) == 2
        for xy in coords:
            assert len(xy) == 2
        self.layer = layer
        self.coords = coords

    def __str__(self):
        leditgrid = 1  # units/nm
        logtext = 'B '
        for xy in self.coords:
            logtext += ('(' + str(int(round(xy[0]*leditgrid))) + ',' +
                        str(int(round(xy[1]*leditgrid))) + ') ')
        return self.layer.name + ': ' + logtext + '\n'

    def export_autogen(self):
        leditgrid = 1  # units/nm
        batchtext = ''
        for xy in self.coords:
            batchtext += (', ' + str(int(round(xy[0]*leditgrid))) +
                          ', ' + str(int(round(xy[1]*leditgrid))))
        batchtext = ('\t\tLBox_New( newCell, ' + self.layer.name + batchtext +
                     ');\n')
        return batchtext

    def makeroundcorneredpolygon(self, chopsize):
        minx = self.coords[0][0]
        miny = self.coords[0][1]
        maxx = minx
        maxy = miny
        for xy in self.coords:
            assert len(xy) == 2
            minx = min(minx, xy[0])
            miny = min(miny, xy[1])
            maxx = max(maxx, xy[0])
            maxy = max(maxy, xy[1])
        coords = [[minx + chopsize * 0, miny + chopsize * 1],
                  [minx + chopsize * 1, miny + chopsize * 0],
                  [maxx - chopsize * 1, miny + chopsize * 0],
                  [maxx - chopsize * 0, miny + chopsize * 1],
                  [maxx - chopsize * 0, maxy - chopsize * 1],
                  [maxx - chopsize * 1, maxy - chopsize * 0],
                  [minx + chopsize * 1, maxy - chopsize * 0],
                  [minx + chopsize * 0, maxy - chopsize * 1],
                  ]
        p = Polygon(self.layer, coords)
        return p

    def __eq__(self, other):
        if not(isinstance(other, Box)):
            return False
        if self.layer != other.layer:
            return False
        if (min(self.coords[0][0], self.coords[1][0]) !=
                min(other.coords[0][0], other.coords[1][0])):
            return False
        if (max(self.coords[0][0], self.coords[1][0]) !=
                max(other.coords[0][0], other.coords[1][0])):
            return False
        if (min(self.coords[0][1], self.coords[1][1]) !=
                min(other.coords[0][1], other.coords[1][1])):
            return False
        if (max(self.coords[0][1], self.coords[1][1]) !=
                max(other.coords[0][1], other.coords[1][1])):
            return False
        return True


class Cell:
    def __init__(self, name, cursor=None, pitch=None):
        if cursor is None:
            cursor = [0, 0]
        if pitch is None:
            pitch = [0, 0]
        self.name = name
        self.polygons = []
        self.boxes = []
        self.cursor = cursor
        self.pitch = pitch

    def __str__(self):
        logtext = 'Cell ' + self.name + ': \n'
        logtext += ('C: (' + str(int(self.cursor[0])) + ', ' +
                    str(int(self.cursor[1])) + ') ')
        logtext += ('P: (' + str(int(self.pitch[0]/10)) + ', ' +
                    str(int(self.pitch[1]/10)) + ') \n')
        for p in self.polygons:
            logtext += str(p)
        for b in self.boxes:
            logtext += str(b)
        return logtext

    def set_cursor(self, cursor):
        self.cursor = cursor

    def addpolygon(self, polygon):
        assert isinstance(polygon, Polygon)
        self.polygons.append(polygon)

    def addbox(self, box):
        assert isinstance(box, Box)
        self.boxes.append(box)

    def isempty(self):
        return len(self.boxes) + len(self.polygons) == 0

    def export_autogen(self, overwrite=False):
        batchtext = '\nnewCell = LCell_Find( activefile, "' + self.name + '");\n'
        if overwrite:
            batchtext += 'if (newCell != NULL) {\n'
            batchtext += '\tLCell_Delete(newCell);\n'
            batchtext += '}\n'
        else:
            batchtext += 'if (newCell == NULL) {\n'
        batchtext += ('\tnewCell  = LCell_New( activefile, "' + self.name +
                      '");\n')
        batchtext += '\tLFile_OpenCell( activefile, "' + self.name + '");\n'

        batchtext += ('\tcoord = LPoint_Set(' + str(int(self.cursor[0])) +
                      ', ' + str(int(self.cursor[1])) + ');\n')
        batchtext += ('\tnewPort = LPort_New( newCell, origin, "cursor", ' +
                      'coord.x, coord.y, coord.x, coord.y);\n')
        batchtext += ('\tLPort_SetTextAlignment( newPort, PORT_TEXT_MIDDLE ' +
                      '| PORT_TEXT_CENTER );\n')
        batchtext += '\tLPort_SetTextSize( newPort, 1 );\n'
        batchtext += ('\tcoord = LPoint_Set(' + str(int(self.pitch[0]/10)) +
                      ', ' + str(int(self.pitch[1]/10)) + ');\n')
        batchtext += ('\tnewPort = LPort_New( newCell, origin, "pitchdiv10",' +
                      ' coord.x, coord.y, coord.x, coord.y);\n')
        batchtext += ('\tLPort_SetTextAlignment( newPort, PORT_TEXT_MIDDLE ' +
                      '| PORT_TEXT_CENTER );\n')
        batchtext += '\tLPort_SetTextSize( newPort, 1 );\n'

        for polygon in self.polygons:
            batchtext += polygon.export_autogen()
        for box in self.boxes:
            batchtext += box.export_autogen()
        # Get the active cell of the active window.
        batchtext += '\tactiveWindow = LWindow_GetVisible();\n'
        batchtext += '\tif( Assigned( activeWindow ) ) {\n'
        batchtext += '\t\tactiveCell = LWindow_GetCell( activeWindow );\n'
        # Check if activeCell equals newCell. (should be)
        batchtext += '\t\tif( activeCell == newCell )\n'
        # Closes activeCell window (except if it is the last one).
        batchtext += '\t\t\tif( LWindow_IsLast(activeWindow) == 0 )\n'
        batchtext += '\t\t\t\tLWindow_Close(activeWindow);\n'
        batchtext += '\t}\n'
        if not overwrite:
            batchtext += '}\n'

        return batchtext

    def __eq__(self, other):
        # Equality does not check for the name being equal (now)
        if not(isinstance(other, Cell)):
            return False
        if self.cursor != other.cursor:
            return False
        if self.pitch != other.pitch:
            return False
        if len(self.polygons) != len(other.polygons):
            return False
        selfpolygons = list(self.polygons)
        otherpolygons = list(other.polygons)
        while len(selfpolygons) > 0:
            selfpol = selfpolygons.pop(0)
            for index in range(len(otherpolygons)):
                if selfpol == otherpolygons[index]:
                    otherpolygons.pop(index)
                    break
            else:
                return False
        assert len(otherpolygons) == 0
        if len(self.boxes) != len(other.boxes):
            return False
        selfboxes = list(self.boxes)
        otherboxes = list(other.boxes)
        while len(selfboxes) > 0:
            selfbox = selfboxes.pop(0)
            for index in range(len(otherboxes)):
                if selfbox == otherboxes[index]:
                    otherboxes.pop(index)
                    break
            else:
                return False
        assert len(otherboxes) == 0
        return True


def header(tech, excelfile):
    batchtext = ("// From vias.py (tech = " + tech + ", xlsfile = " +
                 excelfile + ")\n")
    batchtext += "// Created: " + time.ctime() + ")\n\n"
    batchtext += r"""module autolabel_module
{
#include <stdlib.h>
#include <stdarg.h>
#include <stdio.h>
#include <string.h>
#include <ctype.h>
#include <math.h>

#define EXCLUDE_LEDIT_LEGACY_UPI
#include <ldata.h>
float DESIGN_GRID;
//#include "X:\LEdit\technology\settings.c"

"""
    batchtext += """void layoutbatch()
{
LFile activefile;
LCell newCell;
LCell activeCell;
LPoint Polygon [100];
LWindow activeWindow;
activefile = LFile_GetVisible();
LPoint coord;
LPort newPort;
//LLayer originLayer;
//originLayer = LLayer_Find(activefile,"Origin Layer");

"""
    return batchtext


def readexcel(tech, excelfile):
    wb = openpyxl.load_workbook(filename=excelfile)
    # ws = wb.get_sheet_by_name('vias')
    ws = wb['vias']

    # cell = ws.cell('A1')
    # TOPROW = cell.row
    # LEFTCOL = cell.col_idx
    # I hope min_row and min_column together refer to cell A1
    TOPROW = ws.min_row
    LEFTCOL = ws.min_column
    # print(ws)

    row = TOPROW
    emptyline = 0
    stopAfterSoManyEmptyLines = 10
    while True:
        cell = ws.cell(row=row, column=LEFTCOL)
        if row == TOPROW:
            if cell.value != 'technology name':
                raise Exception('Invalid Excel File')
        else:
            if cell.value == '**General':
                Generalrow = row
            elif cell.value == '**Layersetup':
                Layersetuprow = row
            elif cell.value == '**ViaDRC':
                ViaDRCrow = row
            elif cell.value == '**ConductorDRC':
                ConductorDRCrow = row
            elif cell.value == '**ViaEnclosures':
                ViaEnclosuresrow = row
            elif cell.value == '**End':
                Endrow = row
                break
            if cell.value is None:
                emptyline += 1
            else:
                emptyline = 0
        if emptyline > stopAfterSoManyEmptyLines:
            Endrow = row
            break
        row += 1
    if 'Generalrow' not in locals():
        raise Exception('Invalid Excel File')
    if 'Layersetuprow' not in locals():
        raise Exception('Invalid Excel File')
    if 'ViaDRCrow' not in locals():
        raise Exception('Invalid Excel File')
    if 'ConductorDRCrow' not in locals():
        raise Exception('Invalid Excel File')
    if 'ViaEnclosuresrow' not in locals():
        raise Exception('Invalid Excel File')
    if 'Endrow' not in locals():
        raise Exception('Invalid Excel File')

    techparams = {}
    alllayers = AllLayers()
    # General
    col = LEFTCOL
    while True:
        cell = ws.cell(row=Generalrow + 1, column=col)
        if col == LEFTCOL:
            if cell.value != 'Parameter':
                raise Exception('Invalid Excel File')
        elif col == LEFTCOL + 1:
            if cell.value != 'Value':
                raise Exception('Invalid Excel File')
        else:
            break
        col += 1
    row = Generalrow + 2
    while row < Layersetuprow:
        cell = ws.cell(row=row, column=LEFTCOL)
        paramname = cell.value
        cell = ws.cell(row=row, column=LEFTCOL + 1)
        paramvalue = cell.value
        techparams[paramname] = paramvalue
        row += 1

    # Layersetup
    col = LEFTCOL
    while True:
        cell = ws.cell(row=Layersetuprow + 1, column=col)
        if col == LEFTCOL:
            if cell.value != 'Layername':
                raise Exception('Invalid Excel File')
        elif col == LEFTCOL + 1:
            if cell.value != 'Ledit_layername':
                raise Exception('Invalid Excel File')
        elif col == LEFTCOL + 2:
            if cell.value != 'Ledit_purpose':
                raise Exception('Invalid Excel File')
        elif col == LEFTCOL + 3:
            if cell.value != 'level':
                raise Exception('Invalid Excel File')
        else:
            break
        col += 1
    row = Layersetuprow + 2
    while row < ViaDRCrow:
        cell = ws.cell(row=row, column=LEFTCOL)
        layername = cell.value
        cell = ws.cell(row=row, column=LEFTCOL + 1)
        LEditname = cell.value
        cell = ws.cell(row=row, column=LEFTCOL + 2)
        LEditpurpose = cell.value
        cell = ws.cell(row=row, column=LEFTCOL + 3)
        level = cell.value
        if None in [layername, LEditname, level]:
            raise ViasError('**Layersetup section is having a row with missing crucial information, '
                            '(layername, leditname, level) can not be None for any of these.\nNote: Empty rows are not allowed')
        alllayers.add(layername, LEditname, LEditpurpose, level)
        row += 1

    alllayers.define_toplevel()
    # ViaDimensions
    col = LEFTCOL
    while True:
        cell = ws.cell(row=ViaDRCrow + 1, column=col)
        if col == LEFTCOL:
            if cell.value != 'Layername':
                raise Exception('Invalid Excel File')
        elif col == LEFTCOL + 1:
            if cell.value != 'Width':
                raise Exception('Invalid Excel File')
        elif col == LEFTCOL + 2:
            if cell.value != 'Space':
                raise Exception('Invalid Excel File')
        elif col == LEFTCOL + 3:
            if cell.value is None or cell.value == '':
                logging.warning('Obsoleting Excel File')
                maxviacol = col - 1
                break
            elif cell.value != 'SpaceFarm':
                raise Exception('Invalid Excel File')
        elif col == LEFTCOL + 4:
            if cell.value != 'MaxDimSizeNoFarm':
                raise Exception('Invalid Excel File')
            maxviacol = col
        else:
            break
        col += 1
    row = ViaDRCrow + 2
    while row < ConductorDRCrow:
        cell = ws.cell(row=row, column=LEFTCOL)
        layername = cell.value
        cell = ws.cell(row=row, column=LEFTCOL + 1)
        width = cell.value
        cell = ws.cell(row=row, column=LEFTCOL + 2)
        space = cell.value
        if maxviacol == LEFTCOL + 4:
            cell = ws.cell(row=row, column=LEFTCOL + 3)
            spacefarm = cell.value
            cell = ws.cell(row=row, column=LEFTCOL + 3)
            maxdimsizenofarm = cell.value
        else:
            spacefarm = space
            maxdimsizenofarm = 0
        for layer in alllayers.content:
            if layer.name == layername:
                assert isinstance(layer, Via)
                layer.set_width(width)
                layer.set_space(space)
                layer.set_spacefarm(spacefarm)
                layer.set_maxdimsizenofarm(maxdimsizenofarm)
        row += 1

    # MetalArea
    col = LEFTCOL
    while True:
        cell = ws.cell(row=ConductorDRCrow + 1, column=col)
        if col == LEFTCOL:
            if cell.value != 'Layername':
                raise Exception('Invalid Excel File')
        elif col == LEFTCOL + 1:
            if cell.value != 'Area':
                raise Exception('Invalid Excel File')
        elif col == LEFTCOL + 2:
            if cell.value != 'Width':
                raise Exception('Invalid Excel File')
        elif col == LEFTCOL + 3:
            if cell.value != 'Direction':
                raise Exception('Invalid Excel File')
        else:
            break
        col += 1
    row = ConductorDRCrow + 2
    while row < ViaEnclosuresrow:
        cell = ws.cell(row=row, column=LEFTCOL)
        layername = cell.value
        cell = ws.cell(row=row, column=LEFTCOL + 1)
        area = cell.value
        cell = ws.cell(row=row, column=LEFTCOL + 2)
        width = cell.value
        cell = ws.cell(row=row, column=LEFTCOL + 3)
        direction = cell.value
        for layer in alllayers.content:
            if layer.name == layername:
                assert isinstance(layer, Conductor)
                layer.set_minarea(area)
                layer.set_width(width)
                layer.set_widthrelax(techparams['Grid'])
                layer.set_direction(direction)
                # print(layer)
                # print('layer.width: ' + str(layer.width))
                # print('layer.widthrelax: ' + str(layer.widthrelax))
        row += 1

    # ViaEnclosures
    col = LEFTCOL
    while True:
        cell = ws.cell(row=ViaEnclosuresrow + 1, column=col)
        if col == LEFTCOL:
            if cell.value != 'Via':
                raise Exception('Invalid Excel File')
        elif col == LEFTCOL + 1:
            if cell.value != 'Layername':
                raise Exception('Invalid Excel File')
        elif col == LEFTCOL + 2:
            if cell.value != 'Enc':
                raise Exception('Invalid Excel File')
        elif col == LEFTCOL + 3:
            if cell.value != 'Enc_2opp':
                raise Exception('Invalid Excel File')
        else:
            break
        col += 1
    row = ViaEnclosuresrow + 2
    while row < Endrow:
        cell = ws.cell(row=row, column=LEFTCOL)
        via = cell.value
        cell = ws.cell(row=row, column=LEFTCOL + 1)
        layername = cell.value
        cell = ws.cell(row=row, column=LEFTCOL + 2)
        enc = cell.value
        cell = ws.cell(row=row, column=LEFTCOL + 3)
        enc_2opp = cell.value
        for layer in alllayers.content:
            if layer.name == via:
                assert isinstance(layer, Via)
                layer.add_enclosure(layername, enc, enc_2opp)
        row += 1

    return techparams, alllayers


# def exportc(alllayers):
#    cfiletext = header(tech, excelfile)


def split_via_cond(alllayers):
    vias = [x for x in alllayers.content if 0.25 < x.level % 1 < .75]
    vias.sort()
    conductors = [x for x in alllayers.content if not 0.25 < x.level % 1 < .75]
    conductors.sort()

    return vias, conductors


def calc_fromto(alllayers):
    vias, conductors = split_via_cond(alllayers)
    fromto = []
    for startlayer in conductors:
        if startlayer.level >= 0 and startlayer.level % 1 < .5:
            for endlayer in conductors:
                if (endlayer.level > startlayer.level + 0.5 and not
                        0 < endlayer.level % 1 < .25):
                    fromto.append((startlayer, endlayer))
                    # print(startlayer.name, endlayer.name)

    return vias, conductors, fromto


def calc_stacks(alllayers, MTalias):
    vias, conductors, fromto = calc_fromto(alllayers)

    stacks = {}
    for x in fromto:
        startlayer = x[0]
        endlayer = x[1]
        stackname = startlayer.name + '_' + endlayer.name
        # print(stackname)
        stacks[stackname] = [startlayer]

        thislayer = startlayer
        while thislayer.level < endlayer.level:
            # catch the conductor level higher, but not the conductor level
            # just above a 'real level' (think mimc top plate)
            nextlayerl = [x for x in conductors if thislayer.level + 0.5 <
                          x.level < thislayer.level + 1.5 and not
                          0 < x.level % 1 < .25]
            assert len(nextlayerl) == 1
            nextlayer = nextlayerl[0]
            vial = [x for x in vias if thislayer.level < x.level <
                    nextlayer.level]
            assert len(vial) == 1
            via = vial[0]
            stacks[stackname].extend([via, nextlayer])
            thislayer = nextlayer
        
        if MTalias and endlayer.toplevel:
            aliasstackname = startlayer.name + '_MT'
            stacks[aliasstackname] = list(stacks[stackname])

    # print(stacks)
    return stacks


def draw_via(stackname, stack, techparams, dimx, dimy, viatype, subtype,
             debug=False):
    """dimx x dimy : bottomleft vias on the same centerpoint 0,0
                     all other vias on maxpitch
    viatype:
      'sqr': all conductors are 'square-shaped' and identical in size.
             called square-shaped because Encl_2opp value is considered for all
             sides of via enclosure
        subtype: 0: Conductor layers as small as possible to comply with enclosure rules
                 1: area increased to comply with minimal area rules
                 2: same as 0, but cell edge to cursor distance on halfgrid
                 3: same as 2, but with rounded edges for startlayer
                     only for AA and GC startlayer vias
      'line':
        subtype: 0: minimal enclosure around via
                 1: area increased in length to comply with minimal area rules
                 2: same as 1, but all conductors have the same (widest) width
                 3: minimal enclosure for startlayer, less origin/symmetry constraint on Mx
                     total width/length of cell is only restricted to designgrid/2
                     only for AA and GC startlayer vias
                     condwidth1 and ~2 like condlength1 and ~2
                 4: same as 3, but with rounded edges for startlayer
                     only for AA and GC startlayer vias
      'cross': All conductor enclosures of minimal enclosure around vias,
                  (for each layer separate)
           All conductor lengths (hor/vert) are equal, the max of all hor/vert
            conductor's (min enclosure 2 opposite sides), cell edge to cursor distance on grid
            if minarea rules apply:
              max((minarea of any conductor(hor/vert))/conductorwidth,
                  minwidth of any conductor)
        subtype: 0: minimal enclosure around via total,  length
                 1-8: area increased in length to comply with minimal area rules
                 1: crossing point centered
                 2: crossing point left aligned
                 3: crossing point bottom aligned
                 4: crossing point bottom-left aligned
                 5-8: same as 1-4 respectively, but all conductor widths are
                      equal to the max width of any conductor in the stack going
                      in the same direction
                 9: same as 0, but cell edge to cursor distance on halfgrid
      'big': always 1 by 1, 1umx1um, single via/contact
    subtype : 0 | 1 = center / 2  = left / 3 = bottom / 4 = bottomleft
    general:
    The full width/heigth of the cell is on grid-size
    The cursor is on (-halfgrid | 0),(-halfgrid | 0) and the via will be placed
    initially with the cursor on grid.
    maxpitch is the maximum of (min width + min space) of any of the vias in the
    stack.
    """
    dimstring = str(dimx) + 'x' + str(dimy)
    cellname = stackname + '_' + dimstring + '_' + viatype + '_' + str(subtype)
    # print(cellname)
    farmtype = False
    if 'X' in [dimx, dimy]:
        farmtype = True
        dimx = 2 if dimx == 'X' else dimx
        dimy = 2 if dimy == 'X' else dimy
    for dim in [dimx, dimy]:
        assert dim in [1, 2]

    mangrid = techparams['ManufacturingGrid']
    grid = techparams['Grid']
    halfgrid = grid/2
    assert ((halfgrid / mangrid) % 1 == 0)
    quartergrid = grid/4
    calccursoroffset = (quartergrid/mangrid) % 1 != 0
    assert (viatype in ['sqr', 'line', 'cross', 'big'])
    if debug:
        print(stackname)
    MBBconstraint = grid

    # calculate pitch
    for stacknumber in range(len(stack)):
        layer = stack[stacknumber]
        if layer.isvia():
            if farmtype:
                maxpitch = max(maxpitch, layer.width + layer.space)
                maxpitchfarm = max(maxpitchfarm, layer.width + layer.spacefarm)
            else:
                maxpitch = max(maxpitch, layer.width + layer.space)
        if viatype == 'big':
            pitchgrid = 1000
        else:
            pitchgrid = grid
    if farmtype:
        if math.ceil(maxpitch / pitchgrid) * pitchgrid == math.ceil(maxpitchfarm / pitchgrid) * pitchgrid:
            # no reason to create farmtype cell, return empty cell
            return Cell(cellname)
        else:
            pitch = math.ceil(maxpitchfarm / pitchgrid) * pitchgrid
    else:
        pitch = math.ceil(maxpitch / pitchgrid) * pitchgrid

    # all variables to be calculated: stored in dictionaries with key (layernumber, updown, ishorizontal)
    LTBvia_sqr_A = {}
    LTBvia_sqr_Aarea = {}
    LTBvia_line_L = {}
    LTBvia_line_W = {}

    # Loop through all layers in the stack

    for stacknumber in range(len(stack)):
        layer = stack[stacknumber]
        # from conductor to conductor
        if layer.isvia():
            continue
        horizontal = layer.horizontal
        # grabbing constraints towards vias above and below
        for updown in [-1, 1]:
            # only above AND below if they are not the start or end layer
            if stacknumber + updown not in range(len(stack)):
                continue
            updownlayer = stack[stacknumber + updown]
            assert updownlayer.isvia()
            key = (stacknumber, updown, horizontal)

            LTBvia_sqr_A[key] = max(updownlayer.width + updownlayer.enc[layer.name][1],
                                    layer.width/2)
            LTBvia_sqr_Aarea[key] = max(updownlayer.width + updownlayer.enc[layer.name][1],
                                        layer.width/2, math.sqrt(layer.minarea)/2)

            LTBvia_line_L[key] = max(updownlayer.width + updownlayer.enc[layer.name][1],
                                     layer.width/2)
            LTBvia_line_W[key] = max(updownlayer.width + updownlayer.enc[layer.name][0],
                                     layer.width/2)
            LTBvia_line_area[key] = layer.minarea

    # Create cell and draw conductors
    center00 = [0, 0]
    center11 = [pitch * (dimx - 1), pitch * (dimy - 1)]

    if viatype == 'sqr':
        # calculate A depending on subtype
        if subtype == 0:
            draw_A = math.ceil(max(LTBvia_sqr_A.values())/halfgrid)*halfgrid
            cursor = [draw_A % grid] * 2
        elif subtype == 1:
            draw_A = math.ceil(max(LTBvia_sqr_Aarea.values())/halfgrid)*halfgrid
            cursor = [draw_A % grid] * 2
        elif subtype in [2, 3]:
            # return empty Cell if not from AA/GC to M1
            if stack[0].level > 0.5 or stack[-1].level != 1:
                return Cell(cellname)
            draw_A = math.ceil(max(LTBvia_sqr_A.values())/quartergrid)*quartergrid
            tmpcursor = [draw_A % halfgrid] * 2
            if calccursoroffset and (tmpcursor[0]/mangrid)%1 != 0:
                cursor = [math.ceil((draw_A % halfgrid)/mangrid)*mangrid] * 2
                offset = [cursor[0]-tmpcursor[0], cursor[1]-tmpcursor[1]]
            else:
                cursor = tmpcursor
                offset = [0]*2
            MBBconstraint = halfgrid
        else:
            raise ViasError('subtype not supported (' + cellname + ')')

        # create cell ...
        newcell = Cell(cellname, cursor, [pitch, pitch])
        # ... and add boxes/polygons to it
        for stacknumber in range(len(stack)):
            layer = stack[stacknumber]
            if layer.isconductor():
                if subtype < 3:
                    newcell.addbox(Box(layer, [[center00[0]-draw_A, center00[1]-draw_A],
                                               [center11[0]+draw_A, center11[1]+draw_A]]))
                else:
                    newcell.addpolygon(Box(layer, [[center00[0]-draw_A, center00[1]-draw_A],
                                                   [center11[0]+draw_A, center11[1]+draw_A]]).makeroundcorneredpolygon(halfgrid))
    elif viatype == 'line':
        if subtype == 0:
            draw_L = math.ceil(max(LTBvia_line_L.values())/halfgrid)*halfgrid
            draw_W = math.ceil(max(LTBvia_line_W.values())/halfgrid)*halfgrid
            cursor = [draw_L % grid, draw_W % grid]
        elif subtype == 1:
            for stacknumber in range(len(stack)):
                # LTBvia_line_Larea_tmp =
                draw_L = math.ceil(max(LTBvia_line_L.values())/halfgrid)*halfgrid
            draw_W = math.ceil(max(LTBvia_line_W.values())/halfgrid)*halfgrid
            cursor = [draw_L % grid, draw_W % grid]
        else:
            raise ViasError('subtype not supported (' + cellname + ')')

        for stacknumber in range(len(stack)):
            layer = stack[stacknumber]
            if layer.isconductor():
                if subtype < 2:
                    draw_W = max([LTBvia_line_W[l,u,h] for l,u,h in LTBvia_line_W.keys() if l == stacknumber])
                    newcell.addbox(Box(layer, [[center00[0]-draw_L, center00[1]-draw_W],
                                               [center11[0]+draw_L, center11[1]+draw_W]]))
                else:
                    newcell.addpolygon(Box(layer, [[center00[0]-draw_A, center00[1]-draw_A],
                                                   [center11[0]+draw_A, center11[1]+draw_A]]).makeroundcorneredpolygon(halfgrid))

    # Draw vias
    for stacknumber in range(len(stack)):
        layer = stack[stacknumber]
        if layer.isvia():
            halfviawidth = layer.width / 2
            for y in range(dimy):
                for x in range(dimx):
                    center = [pitch*x, pitch*y]
                    X = [center[0] - halfviawidth, center[0] + halfviawidth]
                    Y = [center[1] - halfviawidth, center[1] + halfviawidth]
                    newcell.addbox(Box(layer,
                                       [[X[0], Y[0]], [X[1], Y[1]]]))



    ###########  Old style, very many complicated, and with bugs
    # VERT = 0 | False
    # HOR = 1 | True
    # properties for each layer in the stack (dict key = index in the stack
    #  list)
    # Width of a layer, viawidth+ minEnclosure
    VE = {}
    # Width of a layer, viawidth+ minEnclosure (respecting win width of layer
    #   itself)
    WE = {}
    # WEhg = WE , but on halfgrid
    WEhg = {}
    WEL = {}
    WEX = {}
    # Length of a layer, viawidth+ minEnclosure@ 2 opposite sides (respecting
    #  win width of layer itself)
    LE2 = {}
    # LE2hg = LE2 , but on halfgrid
    LE2hg = {}
    # Length of a layer, taking min Area into account with width = WE[x]
    # LA = {}
    # Length of a layer, taking min Area into account with width =
    #  maxWE[VERT,HOR]
    # LA2 = {}
    #  min L of square, taking min Area into account with width = WE[x]
    # LSQRA = {}
    #  max of properties possibly split into [VERT,HOR]
    maxpitch = 0
    maxpitchfarm = 0

    maxVE = [0, 0]
    maxWE = [0, 0]
    maxWEhg = [0, 0]
    maxLE2 = [0, 0]
    maxLE2hg = [0, 0]
    # max Length of layer should
    maxLAL = 0
    maxLAX = [0, 0]
    maxLA2L = 0
    maxLA2X = [0, 0]
    maxLSQRA = 0

    # calculate widths and max widths and length for enclosure 2 opposite sides
    for stacknumber in range(len(stack)):
        layer = stack[stacknumber]
        if layer.isconductor():
            tempVE = 0
            tempWE = 0
            tempLE2 = 0
            for updown in [-1, 1]:
                if stacknumber + updown in range(len(stack)):
                    updownlayer = stack[stacknumber + updown]
                    if debug:
                        print('layer: ' + str(layer))
                        print('updownlayer: ' + str(updownlayer))
                        print('updownlayer.enc: ' + str(updownlayer.enc))
                    tempVE = max(tempVE,
                                 updownlayer.width +
                                 updownlayer.enc[layer.name][0]*2)
                    tempWE = max(tempWE,
                                 updownlayer.width +
                                 updownlayer.enc[layer.name][0]*2,
                                 layer.widthrelax)
                    tempLE2 = max(tempLE2,
                                  updownlayer.width +
                                  updownlayer.enc[layer.name][1]*2,
                                  layer.widthrelax)
            # VE: min width based on via + enclosure
            VE[stacknumber] = math.ceil(tempVE / grid) * grid
            # WE: min width based on via + enclosure AND min relaxed width
            WE[stacknumber] = math.ceil(tempWE / grid) * grid
            WEhg[stacknumber] = math.ceil(tempWE / halfgrid) * halfgrid
            LE2[stacknumber] = math.ceil(tempLE2 / grid) * grid
            LE2hg[stacknumber] = math.ceil(tempLE2 / halfgrid) * halfgrid
            # LE2 = math.ceil(tempLE2 / grid) * grid
            maxVE[layer.horizontal] = max(maxVE[layer.horizontal],
                                          VE[stacknumber])
            maxWE[layer.horizontal] = max(maxWE[layer.horizontal],
                                          WE[stacknumber])
            maxWEhg[layer.horizontal] = max(maxWEhg[layer.horizontal],
                                            WEhg[stacknumber])
            maxLE2[layer.horizontal] = max(maxLE2[layer.horizontal],
                                           LE2[stacknumber])
            maxLE2hg[layer.horizontal] = max(maxLE2hg[layer.horizontal],
                                             LE2hg[stacknumber])
            # maxLE2[layer.horizontal] = max(maxLE2[layer.horizontal], LE2)

    # recalculate width to make sure all widths[HV] differ in doublegrid only
    # WE becomes WEL for line, WEX for cross
    for stacknumber in range(len(stack)):
        layer = stack[stacknumber]
        if layer.isconductor():
            if max(maxWE) % doublegrid == 0:
                WEL[stacknumber] = (math.ceil(WE[stacknumber] / doublegrid) *
                                    doublegrid)
            elif max(maxWE) % doublegrid == grid:
                WEL[stacknumber] = math.ceil((WE[stacknumber]-grid) /
                                             doublegrid) * doublegrid + grid
            else:
                raise Exception
            if maxWE[layer.horizontal] % doublegrid == 0:
                WEX[stacknumber] = (math.ceil(WE[stacknumber] / doublegrid) *
                                    doublegrid)
            elif maxWE[layer.horizontal] % doublegrid == grid:
                WEX[stacknumber] = math.ceil((WE[stacknumber]-grid) /
                                             doublegrid) * doublegrid + grid
            else:
                raise Exception

    # calculate lengths for that widths to comply with area constraint
    for stacknumber in range(len(stack)):
        layer = stack[stacknumber]
        if layer.isconductor():

            tempLAL = layer.minarea / (WEL[stacknumber] + pitch*(dimy-1))

            pitchY = pitch*(dimy-1) if layer.horizontal else pitch*(dimx-1)
            pitchX = pitch*(dimx-1) if layer.horizontal else pitch*(dimy-1)
            tempLAX = layer.minarea / (WEX[stacknumber] + pitchY)
            # LA[layernumber] = math.ceil(tempLA / grid) * grid
            LAL = math.ceil(tempLAL / grid) * grid
            LAX = math.ceil(tempLAX / grid) * grid

            tempLSQRA = math.sqrt(max(layer.minarea - (max(maxLE2)+pitchX) *
                                      pitchY - max(maxLE2) * pitchX, 0))
            LSQRA = max(math.ceil(tempLSQRA / grid) * grid, max(maxLE2))

            maxLAL = max(maxLAL, LAL, LE2[stacknumber])
            maxLAX[layer.horizontal] = (
                max(maxLAX[layer.horizontal], LAX, LE2[stacknumber]))
            maxLSQRA = max(maxLSQRA, LSQRA, LE2[stacknumber])

    # then calculate lengths for all layers same width (maxWE[HV?]) to comply
    # with min area constraint
    for stacknumber in range(len(stack)):
        layer = stack[stacknumber]
        if layer.isconductor():
            tempLA2L = layer.minarea / (max(maxWE) + pitch*(dimy-1))
            LA2L = math.ceil(tempLA2L / grid) * grid
            maxLA2L = max(maxLA2L, LA2L, max(maxLE2))

            pitchY = pitch*(dimy-1) if layer.horizontal else pitch*(dimx-1)
            tempLA2X = layer.minarea / (maxWE[layer.horizontal] + pitchY)
            LA2X = math.ceil(tempLA2X / grid) * grid
            maxLA2X[layer.horizontal] = (
                max(maxLA2X[layer.horizontal], LA2X,
                    maxLE2[layer.horizontal]))

    if debug:
        print('stackname: ' + str(stackname))
        print('WE: ' + str(WE))
        print('WEhg: ' + str(WEhg))
        print('WEL: ' + str(WEL))
        print('WEX: ' + str(WEX))
        print('LE2: ' + str(LE2))
        print('LE2hg: ' + str(LE2hg))
        print('maxWE: ' + str(maxWE))
        print('maxWEhg: ' + str(maxWEhg))
        print('maxLE2: ' + str(maxLE2))
        print('maxLE2hg: ' + str(maxLE2hg))
        print('maxLAL: ' + str(maxLAL))
        print('maxLAX: ' + str(maxLAX))
        print('maxLSQRA: ' + str(maxLSQRA))
        print('maxLA2L: ' + str(maxLA2L))
        print('maxLA2X: ' + str(maxLA2X))
    # Calculate some cell parameters
    # cursor = [maxWE[False]/2 % grid, maxWE[True]/2 % grid]
    if viatype == 'sqr':
        if subtype == 0:
            cursor = [max(maxLE2)/2 % grid, max(maxLE2)/2 % grid]
        elif subtype == 1:
            cursor = [maxLSQRA/2 % grid, maxLSQRA/2 % grid]
        elif subtype in [2, 3]:
            cursor = [max(maxLE2hg)/2 % halfgrid, max(maxLE2hg)/2 % halfgrid]
            MBBconstraint = halfgrid
            if subtype == 3 and stack[0].level > 0.5:
                return Cell(cellname)
        else:
            raise ViasError('subtype not supported (' + cellname + ')')
    elif viatype == 'line':
        if subtype == 0:
            halfcondlength = max(maxLE2)/2
        elif subtype == 1:
            halfcondlength = max(maxLAL - pitch * (dimx - 1), max(maxLE2)) / 2
        elif subtype == 2:
            halfcondlength = max(maxLA2L - pitch * (dimx - 1), max(maxLE2)) / 2
        elif subtype in [3, 4]:
            halfcondlength = max(maxLE2hg)/2
            MBBconstraint = halfgrid
            if subtype == 4 and stack[0].level > 0.5:
                return Cell(cellname)
        else:
            raise ViasError('subtype not supported (' + cellname + ')')
        if subtype < 3:
            cursor = [halfcondlength % grid, max(maxWE)/2 % grid]
        else:
            # TODO subtype 3/4
            cursor = [halfcondlength % halfgrid, max(maxWEhg)/2 % halfgrid]
    elif viatype == 'cross':
        cursor = [maxWE[False] / 2 % grid, maxWE[True] / 2 % grid]
    elif viatype == 'big':
        cursor = [1000/2 % grid, 1000/2 % grid]
        pitch = math.ceil(pitch / 1000) * 1000

    # Create cell
    newcell = Cell(cellname, cursor, [pitch, pitch])

    assert cursor in [[0, 0], [MBBconstraint/2, 0], [0, MBBconstraint/2],
                      [MBBconstraint/2, MBBconstraint/2]]

    if cursor != [0, 0] and debug:
        print(newcell.name)

    center00 = [0, 0]
    center11 = [pitch * (dimx - 1), pitch * (dimy - 1)]
    # BBX = [0, 0]
    # BBY = [0, 0]
    for stacknumber in range(len(stack)):
        layer = stack[stacknumber]
        if layer.isvia():
            halfviawidth = layer.width / 2
            for y in range(dimy):
                for x in range(dimx):
                    center = [pitch*x, pitch*y]
                    X = [center[0] - halfviawidth, center[0] + halfviawidth]
                    Y = [center[1] - halfviawidth, center[1] + halfviawidth]
                    newcell.addbox(Box(layer,
                                       [[X[0], Y[0]], [X[1], Y[1]]]))
        if layer.isconductor():
            skipaddbox = False
            if debug:
                print(viatype)
            if viatype == 'sqr':
                # all conducters are of exact same width/length,
                # taking care of te highest of all min enclosures
                # .. and minimal-area-compliant if subtype == 1
                if subtype == 0:
                    halfcondwidth = max(maxLE2) / 2
                elif subtype == 1:
                    halfcondwidth = maxLSQRA / 2
                elif subtype in [2, 3]:
                    halfcondwidth = max(maxLE2hg) / 2
                else:
                    raise ViasError('subtype not supported (' + cellname + ')')
                X = [center00[0] - halfcondwidth, center11[0] + halfcondwidth]
                Y = [center00[1] - halfcondwidth, center11[1] + halfcondwidth]
                if subtype == 3:
                    newcell.addpolygon(Box(layer, [[X[0], Y[0]], [X[1], Y[1]]]).makeroundcorneredpolygon(halfgrid))
                    skipaddbox = True
            elif viatype == 'line':
                halfcondwidth = WEL[stacknumber] / 2
                if subtype == 0:
                    halfcondlength = max(maxLE2) / 2
                elif subtype == 1:
                    halfcondlength = max(maxLAL - pitch * (dimx - 1),
                                         max(maxLE2)) / 2
                elif subtype == 2:
                    halfcondwidth = max(maxWE) / 2
                    halfcondlength = max(maxLA2L - pitch * (dimx - 1),
                                         max(maxLE2)) / 2
                elif subtype in [3, 4]:
                    halfcondlength = max(maxLE2hg) / 2
                    #halfcondlength = WEL[layernumber] / 2
                    halfcondwidth = WEL[stacknumber] / 2
                    halfcondwidth = WEhg[stacknumber] / 2
                    halfcondwidth1 = (math.ceil((halfcondwidth - cursor[1]) /
                                                halfgrid) * halfgrid + cursor[1])
                    halfcondwidth2 = WEhg[stacknumber] - halfcondwidth1
                    if debug:
                        print('halfcondwidth: ' + str(halfcondwidth))
                        print('halfcondwidth1: ' + str(halfcondwidth1))
                        print('halfcondwidth2: ' + str(halfcondwidth2))
                else:
                    raise ViasError('subtype not supported (' + cellname + ')')
                if subtype < 3:
                    X = [center00[0] - halfcondlength,
                         center11[0] + halfcondlength]
                    Y = [center00[1] - halfcondwidth,
                         center11[1] + halfcondwidth]
                else:
                    X = [center00[0] - halfcondlength,
                         center11[0] + halfcondlength]
                    Y = [center00[1] - halfcondwidth1,
                         center11[1] + halfcondwidth2]
                if subtype == 4 and stacknumber == 0:
                    newcell.addpolygon(Box(layer, [[X[0], Y[0]], [X[1], Y[1]]]).makeroundcorneredpolygon(halfgrid))
                    skipaddbox = True
            elif viatype == 'cross':
                halfcondwidth = WEX[stacknumber] / 2
                halfcondlength1_tmp = maxLE2[layer.horizontal] / 2
                halfcondlength1 = (
                        math.ceil((halfcondlength1_tmp -
                                   cursor[not layer.horizontal]) / grid) *
                        grid + cursor[not layer.horizontal])
                if subtype == 0:
                    halfcondlength2 = halfcondlength1
                elif subtype == 9:
                    halfcondlength1 = (
                            math.ceil((halfcondlength1_tmp -
                                       cursor[not layer.horizontal]) / halfgrid) *
                            halfgrid + cursor[not layer.horizontal])
                    halfcondlength2 = halfcondlength1
                    MBBconstraint = halfgrid
                elif 0 < subtype < 9:
                    pitchX = (pitch*(dimx-1) if layer.horizontal else
                              pitch * (dimy-1))
                    if subtype < 5:
                        halfcondlength2_tmp = (
                            max(maxLAX[layer.horizontal] - pitchX -
                                halfcondlength1, halfcondlength1))
                        halfcondlength2 = (
                                math.ceil((halfcondlength2_tmp -
                                           cursor[not layer.horizontal]) /
                                          grid) * grid +
                                cursor[not layer.horizontal])
                    else:
                        halfcondwidth = maxWE[layer.horizontal] / 2
                        halfcondlength2_tmp = (
                            max(maxLA2X[layer.horizontal] - pitchX -
                                halfcondlength1, halfcondlength1))
                        halfcondlength2 = (
                                math.ceil((halfcondlength2_tmp -
                                           cursor[not layer.horizontal]) /
                                          grid) * grid +
                                cursor[not layer.horizontal])
                    if (subtype % 4 == 1 or
                            (layer.horizontal and ((subtype % 4) == 3)) or
                            ((not layer.horizontal) and
                             ((subtype % 4) == 2))):
                        avg_halfcondlength_tmp = (
                                                         halfcondlength1 + halfcondlength2) / 2
                        avg_halfcondlength = (
                                math.ceil((avg_halfcondlength_tmp -
                                           cursor[not layer.horizontal]) /
                                          grid) * grid +
                                cursor[not layer.horizontal])
                        halfcondlength1 = avg_halfcondlength
                        halfcondlength2 = avg_halfcondlength
                else:
                    raise ViasError('subtype not supported (' + cellname + ')')

                if layer.horizontal:
                    X = [center00[0] - halfcondlength1,
                         center11[0] + halfcondlength2]
                    Y = [center00[1] - halfcondwidth,
                         center11[1] + halfcondwidth]
                else:
                    X = [center00[0] - halfcondwidth,
                         center11[0] + halfcondwidth]
                    Y = [center00[1] - halfcondlength1,
                         center11[1] + halfcondlength2]
            elif viatype == 'big':
                halfcondwidth = math.ceil(max(1000, max(maxVE)) / 1000) * 500
                X = [center00[0] - halfcondwidth, center11[0] + halfcondwidth]
                Y = [center00[1] - halfcondwidth, center11[1] + halfcondwidth]
            if not skipaddbox:
                newcell.addbox(Box(layer, [[X[0], Y[0]], [X[1], Y[1]]]))
            try:
                if viatype != 'big':
                    assert min(Y[1] - Y[0], X[1] - X[0]) >= layer.widthrelax

                assert (X[0] - cursor[0]) % MBBconstraint == 0
                assert (X[1] - cursor[0]) % MBBconstraint == 0
                assert (Y[0] - cursor[1]) % MBBconstraint == 0
                assert (Y[1] - cursor[1]) % MBBconstraint == 0
            except AssertionError:
                logging.error("Assertion Error on cell: " + newcell.name)
                # raise
            # BBX=[math.floor((min(BBX[0], X[0]) - cursor[0])/grid) * grid +
            #  cursor[0], math.ceil((max(BBX[1], X[1]) - cursor[0])/grid) *
            #  grid + cursor[0]]
            # BBY=[math.floor((min(BBY[0], Y[0]) - cursor[1])/grid) *grid +
            #  cursor[1], math.ceil((max(BBY[1], Y[1]) - cursor[1])/grid) *
            #  grid + cursor[1]]
    # celloutline = Layer('celloutline', 'Cell Outline Layer', None, -1)
    # newcell.addpolygon(Polygon(celloutline, [[BBX[0],BBY[0]],[BBX[1],BBY[0]],
    #  [BBX[1],BBY[1]],[BBX[0],BBY[1]]]))

    return newcell


def draw_vias(stacks, techparams, keepallvariants=False):
    # Draw the following via combinations:
    # draw_1x1_sqr()
    # +------+
    # |      |
    # |  XX  |
    # |      |
    # +------+
    # draw_1x1_sqr0()
    # /------\
    # |      |
    # |  XX  |
    # |      |
    # \------/

    # draw_1x1_line()
    # /------\
    # +------+
    # |  XX  |
    # +------+
    # \------/

    # draw_1x1_linearea()
    # +----------+
    # |    XX    |
    # +----------+

    # draw_1x1_cross()
    #     +--+
    #     |  |
    # +----------+
    # |    XX    |
    # +----------+
    #     |  |
    #     +--+

    # draw_1x1_cross2()
    #   +--+
    #   |  |
    # +----------+
    # |  XX      |
    # +----------+
    #   |  |
    #   +--+

    # draw_1x1_cross3()
    #     +--+
    #     |  |
    #     |  |
    # +----------+
    # |    XX    |
    # +----------+
    #     +--+

    # draw_1x1_cross4()
    #   +--+
    #   |  |
    #   |  |
    # +----------+
    # |  XX      |
    # +----------+
    #   +--+

    # draw_1x1_big()
    # +----------+
    # |          |
    # |          |
    # |    XX    |
    # |          |
    # |          |
    # +----------+

    # for each stack, there are 3 lengths defined
    # max of each via in stack, ceil grid (or grid/2)
    # minenc = half of the viawidth + min enclosure
    # oppenc = half of the viawidth + min enclosure on 2 opposite sides
    # areaenc = half of the viawidth + minarea of conductor
    # areaenc_asym = minarea of conductor asymmetric because of cross
    stacknames = list(stacks.keys())
    stacknames.sort()
    cells = []
    for stackname in stacknames:
        if stackname in ['AA_M4']:
            # 'AA_M1', 'M1_M2',
            debug = True
        else:
            debug = False
        # debug = False
        stack = stacks[stackname]

        for dimx, dimy in [(1, 1), (2, 1), ('X', 1), (1, 2), (1, 'X'), (2, 2), ('X', 'X')]:
            # for dimx, dimy in [(1, 1), (2, 1), (1, 2), (2, 2)]:
            # for dimx, dimy in [(1, 1), (1, 2), (1, 'X'), (2, 1), ('X', 1), (2, 2), ('X', 'X')]:
            cellsdims = []
            cellsdiml = []
            cellsdimc = []
            cellsdimb = []
            no_subtypes = [4, 5, 10]   # [2,3,9] old lib (for debug and check equality)
            for subtype in range(no_subtypes[0]):
                cellsdims.append(draw_via(stackname, stack, techparams, dimx,
                                          dimy, 'sqr', subtype, debug))
            for subtype in range(no_subtypes[1]):
                cellsdiml.append(draw_via(stackname, stack, techparams, dimx,
                                          dimy, 'line', subtype, debug))
            for subtype in range(no_subtypes[2]):
                cellsdimc.append(draw_via(stackname, stack, techparams, dimx,
                                          dimy, 'cross', subtype, debug))
            cellsdimb.append(draw_via(stackname, stack, techparams, dimx,
                                      dimy, 'big', 0, debug))

            # keep at least 1 of each kind
            for cellsbytype in [cellsdims, cellsdiml, cellsdimc, cellsdimb]:
                # delete duplicates and empty cells
                for i in range(len(cellsbytype) - 1, -1, -1):
                    if cellsbytype[i].isempty():
                        p = cellsbytype.pop(i)
                        if debug:
                            print('empty: ' + str(p.name))
                    else:
                        if not keepallvariants:
                            for j in range(i):
                                if cellsbytype[i] == cellsbytype[j]:
                                    p = cellsbytype.pop(i)
                                    # if debug:
                                    if True:
                                        if debug:
                                            print('duplicate: ' + str(p.name) +
                                                  '   keep: ' +
                                                  str(cellsbytype[j].name))
                                    break
                cells.extend(cellsbytype)

    return cells


def viagenwrite(tech, outfilepath, outfile, filenr, batch, backup):
    # exactly one of both (outfile,outfilepath) is None
    if outfilepath is not None:
        thisoutfile = outfilepath + tech + '_vias_' + str(filenr) + '.c'
    else:
        thisoutfile = outfile + ('_' + str(filenr) if filenr > 1 else '')
    print(thisoutfile)
    general.write(thisoutfile, batch, backup)


def vialogwrite(tech, outfilepath, outfile, batch, backup):
    # exactly one of both (outfile,outfilepath) is None
    if outfilepath is not None:
        thisoutfile = outfilepath + tech + '_vias.log'
    else:
        thisoutfile = outfile
    print(thisoutfile)
    general.write(thisoutfile, batch, backup)


def generate(tech=None, excelfile=None, overwrite=False, outfile=None,
             logfile=None, keepallvariants=False, backup=True, MTalias=False):
    if tech is None:
        technames = []
        for file in os.listdir(LTBsettings.viasexcelfilepath()):
            if file.endswith('.xlsx'):
                technames.append(file[:-5])
                print(file[:-5])
    else:
        technames = [tech]

    if outfile is None:
        outfilepath = LTBsettings.laygenfilepath()
    else:
        outfilepath = None

    if logfile is None:
        logfilepath = LTBsettings.viasexcelfilepath()
    else:
        logfilepath = None

    # exactly one of both (outfile,outfilepath) is None

    for tech in technames:
        if excelfile is None:
            thisexcelfile = LTBsettings.viasexcelfile(tech)
        else:
            thisexcelfile = excelfile

        print(thisexcelfile)
        techparams, alllayers = readexcel(tech, thisexcelfile)

        stacks = calc_stacks(alllayers, MTalias)

        cells = draw_vias(stacks, techparams, keepallvariants)

        batchbegin = header(tech, thisexcelfile)
        batchbegin += alllayers.export_autogen()
        batchbegin += '\n'

        batchend = '}\n'
        batchend += '}\n\n'
        batchend += 'layoutbatch();\n'

        print('Number of cells: ' + str(len(cells)))
        batchtext = ''
        logtext = ''
        filenr = 1
        for cellnr in range(len(cells)):
            cell = cells[cellnr]
            batchtext += cell.export_autogen(overwrite)
            logtext += str(cell)
            if cellnr % 500 == 499 or cellnr == len(cells)-1:
                viagenwrite(tech, outfilepath, outfile, filenr,
                            batchbegin + batchtext + batchend, backup)
                filenr += 1
                batchtext = ''
        vialogwrite(tech, logfilepath, logfile, logtext, backup)


def argparse_setup(subparsers):
    parser_gen = subparsers.add_parser(
            'viagenerate', help='Creates L-Edit C-file to generate vias')
    parser_gen.add_argument(
            '-t', '--technology', default=None, help='the TECHNOLOGY name ' +
            '(default: loops over all excel files in ' +
            LTBsettings.viasexcelfilepath() + ')')
    parser_gen.add_argument(
            '-x', '--excelfile', default=None, help='the path to the excel ' +
            'file (default: ' + LTBsettings.viasexcelfilepath() +
            'TECHNOLOGY.xlsx)')
    parser_gen.add_argument(
            '-w', '--overwrite', default=False, action='store_true',
            help='Overwrite existing cells in the design')
    parser_gen.add_argument(
            '-o', '--outfile', default=None, help='the path to the output ' +
            'file (default: ' + LTBsettings.laygenfilepath() +
            'TECHNOLOGY_vias_#.c)')
    parser_gen.add_argument(
            '-l', '--logfile', default=None, help='the path to the log file ' +
            '(default: ' + LTBsettings.viasexcelfilepath() +
            'TECHNOLOGY_vias.log)')
    parser_gen.add_argument(
            '-k', '--keepallvariants', default=False, action='store_true',
            help='Keep all variants, do not remove duplicates of other cells.')
    parser_gen.add_argument(
            '--nobackup', dest='backup', default=True, action='store_false',
            help='Avoids creation of backup files of previous output files.')
    parser_gen.add_argument(
            '--MTalias', default=False, action='store_true',
            help='Creates MT alias for the endlayer.')


def argparse_eval(args):
    # make a dict of args
    dictargs = vars(args)
    funcdict = {'viagenerate': (generate,
                                [dictargs.get('technology'),
                                 dictargs.get('excelfile'),
                                 dictargs.get('overwrite'),
                                 dictargs.get('outfile'),
                                 dictargs.get('logfile'),
                                 dictargs.get('keepallvariants'),
                                 dictargs.get('backup'),
                                 dictargs.get('MTalias')])
                }
    return funcdict


if __name__ == "__main__":
    general.logsetup()
    general.myargparse(argparse_setup, argparse_eval, 'v20240909')
