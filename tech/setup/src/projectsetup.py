# -*- coding: utf-8 -*-
"""
Created on Fri Mar 25 11:06:32 2016

@author: Koen
"""
import openpyxl
# import logging      # in case you want to add extra logging
import general
import LTBsettings


def projectsetup(excelfile=None, outfilename=None, backup=False):
    if excelfile is None:
        excelfile = LTBsettings.projectsexcelfile()
    if outfilename is None:
        outfilename = LTBsettings.defaultprojectsettings()
    wb = openpyxl.load_workbook(filename=excelfile)
    # ws = wb.get_sheet_by_name('Projects')
    ws = wb['Projects']

    # cell = ws.cell('A1')
    # TOPROW = cell.row
    # LEFTCOL = cell.col_idx
    # I hope min_row and min_column together refer to cell A1
    TOPROW = ws.min_row
    LEFTCOL = ws.min_column
    # print(ws)

    col = LEFTCOL
    projectdictlist = []
    while True:
        # for i in range(10):
        cell = ws.cell(row=TOPROW, column=col)
        if col == LEFTCOL:
            if cell.value != 'projectname':
                raise Exception('Invalid Excel File')
        else:
            if cell.value is not None:
                projname = cell.value
                projectdictlist.append({'name': projname, 'col': col})
            else:
                break
        col += 1

    outputtext = ('** Python project settings export from source: ' +
                  excelfile + '\n\n')
    outputtext += ("** DO NOT MODIFY FILE (unless you know very well what " +
                   "you're doing.)\n\n")

    for projectdict in projectdictlist:

        projectname = projectdict['name']
        projectcol = projectdict['col']
        # columnstowrite = [LEFTCOL, projectcol]
        row = TOPROW
        emptyline = 0
        stopAfterSoManyEmptyLines = 10
        while True:
            settingname = ws.cell(row=row, column=1).value
            settingname = (str(settingname) if settingname is not None
                           else settingname)
            value = ws.cell(row=row, column=projectcol).value
            value = str(value) if value is not None else value
            if settingname is None:
                emptyline += 1
            else:
                if value is None:
                    outputtext += ('* MISSING VALUE (or obsolete ' +
                                   'settingname) for ' + settingname + '\n')
                else:
                    outputtext += (projectname + '.' + settingname + ' = ' +
                                   str(value) + '\n')
                emptyline = 0

            if emptyline > stopAfterSoManyEmptyLines:
                outputtext += '\n\n'
                break
            row += 1

    print(outfilename)
    general.write(outfilename, outputtext, backup)


def argparse_setup(subparsers):
    parser_proj = subparsers.add_parser(
            'projects', help='Creates projects.ini file for all projects')
    parser_proj.add_argument(
            '-o', '--outfile', default=None, help='the path to the output ' +
            'file (default: ' + LTBsettings.defaultprojectsettings() + ')')
    parser_proj.add_argument(
            '-x', '--excelfile', default=None, help='the path to the excel ' +
            'file (default: ' + LTBsettings.projectsexcelfile() + ')')
    parser_proj.add_argument(
            '--nobackup', dest='backup', default=True, action='store_false',
            help='Avoids creation of backup files of previous output files.')


def argparse_eval(args):
    # make a dict of args
    dictargs = vars(args)
    funcdict = {'projects': (projectsetup,
                             [dictargs.get('excelfile'),
                              dictargs.get('outfile'),
                              dictargs.get('backup')])
                }
    return funcdict


if __name__ == "__main__":
    general.logsetup()
    general.myargparse(argparse_setup, argparse_eval, 'v20221005')
