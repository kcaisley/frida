# -*- coding: utf-8 -*-
"""
Created on Fri Mar 25 11:06:32 2016

@author: Koen
"""
import os
# consider xlwings instead of openpyxl if formulas are to be evaluated
import openpyxl
import json
import logging

import general
import LTBsettings

import vias


def v2common(tech=None, excelfile=None, backup=False):
    if tech is None:
        tech = 'all'
    if excelfile is None:
        excelfile = LTBsettings.v2commonexcelfile()

    wb = openpyxl.load_workbook(filename=excelfile)
    # ws = wb.get_sheet_by_name('Technologies')
    ws = wb['Technologies']

    # cell = ws.cell('A1')
    # TOPROW = cell.row
    # LEFTCOL = cell.col_idx
    # I hope min_row and min_column together refer to cell A1
    TOPROW = ws.min_row
    LEFTCOL = ws.min_column
    # print(ws)

    row = TOPROW
    emptyline = 0
    stopAfterSoManyEmptyLines = 10
    v2commonfilenamerow = None
    v2common_linux1row = None
    v2common_linux2row = None
    while True:
    # for i in range(10):
        cell = ws.cell(row = row, column=LEFTCOL)
        if row == TOPROW:
            if cell.value != 'technology name':
                raise Exception('Invalid Excel File')
        else:
            if cell.value == 'v2.0_common.sp':
                v2commonfilenamerow = row
            elif cell.value == 'v2common_linux1':
                v2common_linux1row = row
            elif cell.value == 'v2common_linux2':
                v2common_linux2row = row
            if cell.value == None:
                emptyline += 1
            else:
                emptyline = 0
        if emptyline > stopAfterSoManyEmptyLines:
            break
        row += 1
    # print(v2commonfilenamerow)

    col = LEFTCOL
    techdictlist = []
    while True:
    #for i in range(10):
        cell = ws.cell(row = TOPROW, column=col)
        if col == LEFTCOL:
            if cell.value != 'technology name':
                raise Exception('Invalid Excel File')
        else:
            if cell.value == None:
                break
            if tech == 'all' or tech == cell.value:
                techname = cell.value
                if v2commonfilenamerow is not None:
                    cell = ws.cell(row = v2commonfilenamerow, column=col)
                    winpath = cell.value
                else:
                    winpath = None
                if v2common_linux1row is not None:
                    cell = ws.cell(row = v2common_linux1row, column=col)
                    lin1path = cell.value
                else:
                    lin1path = None
                if v2common_linux2row is not None:
                    cell = ws.cell(row = v2common_linux2row, column=col)
                    lin2path = cell.value
                else:
                    lin2path = None
                techdictlist.append({'name': techname, 'col': col, 'winpath' : winpath, 'lin1path' : lin1path, 'lin2path' : lin2path})
        #print('col ' + str(col) + ': ' + cell.value)
        col += 1
    #print(techdictlist)

    #ws = wb.get_sheet_by_name('v2.0_common')
    ws = wb['v2.0_common']
    #print(ws)

    for techdict in techdictlist:

        col = LEFTCOL
        while True:
            cell = ws.cell(row = TOPROW, column=col)
            #print('col ' + str(col) + ': ' + str(cell.value))
            if cell.value == techdict['name']:
                v2commonvaluecol = col
                break
            if col > 10:
                raise Exception('Technology file not complete or larger than expected.\n' + techdict['name'] + ' not found in first row of v2.0_common, searched up to ' + str(col) + ' columns.')
                break
            col +=1

        commentcol = LEFTCOL
        paramcol = LEFTCOL+1
        #columnstowrite = [LEFTCOL, LEFTCOL+1, v2commonvaluecol]
        row = TOPROW+1
        outputtext = '** Python v2.0_common export from source: ' + excelfile + '\n\n'
        outputtext += "** DO NOT MODIFY FILE (unless you know very well what you're doing.)\n\n"
        emptyline = 0
        stopAfterSoManyEmptyLines = 10
        while True:
            comment = ws.cell(row = row, column=commentcol).value
            comment = str(comment) if comment is not None else comment
            param = ws.cell(row = row, column=paramcol).value
            param = str(param) if param is not None else param
            value = ws.cell(row = row, column=v2commonvaluecol).value
            value = str(value) if value is not None else value
            if comment is None and param is None:
                linetext = ''
                emptyline += 1
            else:
                if comment is None and param is not None:
                    if value is None:
                        linetext = '* MISSING VALUE (or obsolete param)'
                    else:
                        linetext = ''
                    linetext += '.param ' + param + ' = '
                    linetext += value if value is not None else ''
                elif comment is not None and param is None:
                    linetext = '*' + comment
                elif comment is not None and param is not None:
                    linetext = '*' + comment + '\n'
                    if value is None:
                        linetext += '* '
                    linetext += '.param ' + param + ' = '
                    try:
                        linetext += value if value is not None else ''
                    except:
                        print(linetext)
                        print(value)
                emptyline = 0

            outputtext += linetext + '\n'
            if emptyline > stopAfterSoManyEmptyLines:
                break
            row += 1

        outpaths = [techdict['winpath'],techdict['lin1path'],techdict['lin2path']]
        for outpath in outpaths:
            if outpath is None:
                continue
            outfilename = outpath + r'\v2.0_common.sp'
            cut = -1 * (stopAfterSoManyEmptyLines - 3)
            try:
                general.write(outfilename, outputtext[:cut], backup)
            except:
                if backup:
                    logtext = 'Failed: renaming of ' + outfilename + ' to ' + outfilename + '.bak_##\n'
                else:
                    logtext = 'Failed: removal of ' + outfilename + '\n'
                logtext += '        or creation of its filepath or file. See log for more details'
                print(logtext)
                general.error_log(logtext)
                general.error_log()
                continue

            print('Ok: ' + outfilename)


def tech2layoutparams(excelfile = None, outfilename = None, backup = False):
    if True:
        tech = 'all'
    if excelfile is None:
        excelfile = LTBsettings.tech2layoutparamsexcelfile()
    if outfilename is None:
        outfilename = LTBsettings.defaulttech2layoutparams()

    wb = openpyxl.load_workbook(filename = excelfile)
    #ws = wb.get_sheet_by_name('Technologies')
    ws = wb['Technologies']

    #cell = ws.cell('A1')
    #TOPROW = cell.row
    #LEFTCOL = cell.col_idx
    # I hope min_row and min_column together refer to cell A1
    TOPROW = ws.min_row
    LEFTCOL = ws.min_column
    #print(ws)

    row = TOPROW
    emptyline = 0
    stopAfterSoManyEmptyLines = 10
    while True:
    #for i in range(10):
        cell = ws.cell(row = row, column=LEFTCOL)
        if row == TOPROW:
            if cell.value != 'technology name':
                raise Exception('Invalid Excel File')
        else:
            if cell.value == None:
                emptyline += 1
            else:
                emptyline = 0
        if emptyline > stopAfterSoManyEmptyLines:
            break
        row += 1

    col = LEFTCOL
    techdictlist = []
    while True:
    #for i in range(10):
        cell = ws.cell(row = TOPROW, column=col)
        if col == LEFTCOL:
            if cell.value != 'technology name':
                raise Exception('Invalid Excel File')
        else:
            if cell.value == None:
                break
            if tech == 'all' or tech == cell.value:
                techname = cell.value
                techdictlist.append({'name': techname})
        #print('col ' + str(col) + ': ' + cell.value)
        col += 1
    #print(techdictlist)

    outputtext = ('/*\n** Python techlayers export from source: ' +
                  excelfile + '\n\n')
    outputtext += ("** DO NOT MODIFY FILE (unless you know very well what " +
                   "you're doing.)\n*/\n\n")

    # ws = wb.get_sheet_by_name('techlayers')
    ws = wb['techlayers']
    # print(ws)

    outputtext += """LLayer tech2layer(char* layername)
{
	LFile activefile;
	char tech[128];
	char project[128];
	LEdittech2tech_project(tech,project,1);

	LLayer requestedLayer;

	activefile = LFile_GetVisible();
	requestedLayer = NULL;

"""
    commonlayer = """void commonlayer(int thislayer, char* requestedLayer){
"""
    commonlayerint=0

    first = True
    for techdict in techdictlist:

        col = LEFTCOL
        while True:
            cell = ws.cell(row = TOPROW, column=col)
            #print('col ' + str(col) + ': ' + str(cell.value))
            if cell.value == techdict['name']:
                layernamecol = col
                break
            if col > 10:
                raise Exception('Technology file not complete or larger than expected.\n' + techdict['name'] + ' not found in first row of v2.0_common, searched up to ' + str(col) + ' columns.')
                break
            col +=1

        outputtext += '\t'
        if not first:
            outputtext += 'else '
        #move to end of loop
        #else:
        #    first = False
        outputtext += 'if (strcmp(tech, "' + techdict['name'] + '") == 0) \n\t{\n'

        commentcol = LEFTCOL
        generalnamecol = LEFTCOL+1
        #columnstowrite = [LEFTCOL, LEFTCOL+1, layernamecol]
        row = TOPROW+1
        emptyline = 0
        stopAfterSoManyEmptyLines = 10
        while True:
            comment = ws.cell(row = row, column=commentcol).value
            comment = str(comment) if comment is not None else comment
            param = ws.cell(row = row, column=generalnamecol).value
            param = str(param) if param is not None else param
            value = ws.cell(row = row, column=layernamecol).value
            value = str(value) if value is not None else value
            linetext = ''
            if comment is None and param is None:
                emptyline += 1
            else:
                if comment is not None:
                    linetext = '\t\t//' + comment
                if param is not None:
                    linetext += '\t\tif (strcmp( layername, "' + param +'")==0) {\n'
                    if first:
                        commonlayer += '''	if (thislayer ==''' + str(commonlayerint) + '''){
		sprintf(requestedLayer, "''' + param + '''");}
'''
                        commonlayerint += 1

                    if value is None:
                        linetext += '\t\t\t//MISSING VALUE\n\t\t\trequestedLayer = NULL; }'
                    elif value == "NULL":
                        linetext += '\t\t\trequestedLayer = NULL; }'
                    elif ":" in value:
                        layer, purpose = value.split(':',1)
                        linetext += '\t\t\trequestedLayer = LLayer_FindByNames(activefile,"' + layer + '", "' + purpose + '"); }'
                    else:
                        linetext += '\t\t\trequestedLayer = LLayer_Find(activefile,"' + value + '"); }'
                emptyline = 0

            outputtext += linetext + '\n'
            if emptyline == stopAfterSoManyEmptyLines:
                outputtext = outputtext[:-stopAfterSoManyEmptyLines] + '\t}\n'
                break
            row += 1

        if first:
            commonlayer += """
	else {
		requestedLayer = NULL;}
	return requestedLayer;
}

int numberofcommonlayers(void){
	return """ + str(commonlayerint) + """;
}

"""
        first = False

    outputtext += """
	return requestedLayer;
}

"""

    outputtext += commonlayer



    #ws = wb.get_sheet_by_name('techglobals')
    ws = wb['techglobals']
    #print(ws)

    outputtext += """void tech2grid(void)
{
	char tech[128];
	char project[128];
	LEdittech2tech_project(tech,project,1);

"""
    first = True
    for techdict in techdictlist:

        col = LEFTCOL
        while True:
            cell = ws.cell(row=TOPROW, column=col)
            # print('col ' + str(col) + ': ' + str(cell.value))
            if cell.value == techdict['name']:
                layernamecol = col
                break
            if col > 10:
                raise Exception('Technology file not complete or larger than' +
                                ' expected.\n' + techdict['name'] + ' not ' +
                                'found in first row of v2.0_common, searched' +
                                ' up to ' + str(col) + ' columns.')
                break
            col += 1

        outputtext += '\t'
        if not first:
            outputtext += 'else '
        else:
            first = False
        outputtext += ('if (strcmp(tech, "' + techdict['name'] +
                       '") == 0) \n\t{\n')

        commentcol = LEFTCOL
        globalname = LEFTCOL+1
        # columnstowrite = [LEFTCOL, LEFTCOL+1, layernamecol]
        row = TOPROW+1
        emptyline = 0
        stopAfterSoManyEmptyLines = 10
        while True:
            comment = ws.cell(row=row, column=commentcol).value
            comment = str(comment) if comment is not None else comment
            param = ws.cell(row=row, column=globalname).value
            param = str(param) if param is not None else param
            value = ws.cell(row=row, column=layernamecol).value
            value = str(value) if value is not None else value
            linetext = ''
            if comment is None and param is None:
                emptyline += 1
            else:
                if comment is not None:
                    linetext += '\t\t//' + comment
                if param is not None:
                    if value is None:
                        linetext += ('\t\t//MISSING VALUE\n\t\t' + param +
                                     ' = NULL;')
                    else:
                        linetext += '\t\t' + param + ' = ' + value + ';'
                emptyline = 0

            outputtext += linetext + '\n'
            if emptyline == stopAfterSoManyEmptyLines:
                outputtext = outputtext[:-stopAfterSoManyEmptyLines] + '\t}\n'
                break
            row += 1
    outputtext += """
}

"""

    try:
        general.write(outfilename, outputtext, backup)
    except:
        if backup:
            logtext = ('Failed: renaming of ' + outfilename + ' to ' +
                       outfilename + '.bak_##\n')
        else:
            logtext = 'Failed: removal of ' + outfilename + '\n'
        logtext += ('        or creation of its filepath or file. ' +
                    'See log for more details')
        print(logtext)
        general.error_log(logtext)
        general.error_log()

    print('Ok: ' + outfilename)


def gdssheet(excelfile=None, outfilename=None, backup=False):
    if True:
        tech = 'all'
    if excelfile is None:
        excelfile = LTBsettings.tech2layoutparamsexcelfile()
    if outfilename is None:
        outfilename = LTBsettings.defaultgdssheet()

    wb = openpyxl.load_workbook(filename=excelfile)
    # ws = wb.get_sheet_by_name('Technologies')
    ws = wb['Technologies']

    # cell = ws.cell('A1')
    # TOPROW = cell.row
    # LEFTCOL = cell.col_idx
    # I hope min_row and min_column together refer to cell A1
    TOPROW = ws.min_row
    LEFTCOL = ws.min_column
    # print(ws)

    row = TOPROW
    emptyline = 0
    stopAfterSoManyEmptyLines = 10

    # for i in range(10):
    while True:
        cell = ws.cell(row=row, column=LEFTCOL)
        if row == TOPROW:
            if cell.value != 'technology name':
                raise Exception('Invalid Excel File')
        else:
            if cell.value is None:
                emptyline += 1
            else:
                emptyline = 0
        if emptyline > stopAfterSoManyEmptyLines:
            break
        row += 1

    col = LEFTCOL
    techdictlist = []

    # for i in range(10):
    while True:
        cell = ws.cell(row=TOPROW, column=col)
        if col == LEFTCOL:
            if cell.value != 'technology name':
                raise Exception('Invalid Excel File')
        else:
            if cell.value is None:
                break
            if tech == 'all' or tech == cell.value:
                techname = cell.value
                techdictlist.append({'name': techname})
        # print('col ' + str(col) + ': ' + cell.value)
        col += 1
    # print(techdictlist)

    # ws = wb.get_sheet_by_name('techlayers')
    ws = wb['GDSsheet']
    # print(ws)

    for techdict in techdictlist:

        col = LEFTCOL
        while True:
            cell = ws.cell(row=TOPROW, column=col)
            # print('col ' + str(col) + ': ' + str(cell.value))
            if cell.value == techdict['name']:
                gdsnumcol = col
                gdsdtcol = col + 1
                break
            if col > (len(techdictlist)*2)+2:
                raise Exception('Technology file not complete or larger than' +
                                ' expected.\n' + techdict['name'] +
                                ' not found in first row of ' + ws.title + ', ' +
                                'searched up to ' + str(col) + ' columns.')
                break
            col += 1

        commentcol = LEFTCOL
        generalnamecol = LEFTCOL+1
        # columnstowrite = [LEFTCOL, LEFTCOL+1, layernamecol]
        row = TOPROW+1
        emptyline = 0
        stopAfterSoManyEmptyLines = 10
        while True:
            comment = ws.cell(row=row, column=commentcol).value
            comment = comment if isinstance(comment, str) else None
            param = ws.cell(row=row, column=generalnamecol).value
            param = param if isinstance(param, str) else None
            gdsnum = ws.cell(row=row, column=gdsnumcol).value
            gdsnum = gdsnum if isinstance(gdsnum, int) else None
            gdsdt = ws.cell(row=row, column=gdsdtcol).value
            gdsdt = gdsdt if isinstance(gdsdt, int) else None
            linetext = ''
            if comment is None and param is None:
                emptyline += 1
            else:
                if None not in [param, gdsnum, gdsdt]:
                    techdict[param] = str([gdsnum, gdsdt])[1:-1]
                    techdict[str([gdsnum, gdsdt])[1:-1]] = param

                emptyline = 0

            if emptyline == stopAfterSoManyEmptyLines:
                break
            row += 1

    general.prepare_write(outfilename, backup)
    with open(outfilename, 'w') as fp:
        json.dump(techdictlist, fp, indent=2)

    print('Ok: ' + outfilename)


def deltdo(path, dryrun, sizesummary=False):
    dryruntext = ''
    if dryrun:
        dryruntext = '-- dryrun -- '
    size = 0
    fullpath = 'S:\\projects\\' + path
    for fileordir in os.listdir(fullpath):
        fullfileordir = fullpath + os.sep + fileordir
        if os.path.isdir(fullfileordir):
            size += deltdo(fullfileordir, dryrun)
        elif os.path.isfile(fullfileordir):
            if fullfileordir.endswith('.tdo'):
                thissize = os.path.getsize(fullfileordir)
                size += thissize
                logtext = (dryruntext + 'DELTDO remove: ' + fullfileordir +
                           ' (' + str(round(thissize / 1024 / 1024, 2)) +
                           ' Mb)')
                print(logtext)
                general.error_log(logtext)
                if not dryrun:
                    os.remove(fullfileordir)

    if sizesummary:
        print(dryruntext + 'total space created: ' +
              str(round(size / 1024 / 1024, 2)) + ' Mb')

    return size


def argparse_setup(subparsers):
    parser_v2c = subparsers.add_parser(
            'v2common',
            help=('Creates v2.0_common.sp files for all or a specified ' +
                  'technology'))
    parser_v2c.add_argument('-t', '--technology', default=None,
                            help=('the TECHNOLOGY name (default: loops over ' +
                                  'all technologies defined in EXCELFILE)'))
    parser_v2c.add_argument('-x', '--excelfile', default=None,
                            help=('the path to the excel file (default: ' +
                                  LTBsettings.v2commonexcelfile() + ')'))
    parser_v2c.add_argument('--nobackup', dest='backup', default=True,
                            action='store_false',
                            help=('Avoids creation of backup files of ' +
                                  'previous output files.'))

    parser_t2l = subparsers.add_parser(
            'tech2layoutparams',
            help='Creates the layout parameter file for all technologies')
    parser_t2l.add_argument('-o', '--outfile', default=None,
                            help=('the path to the output file (default: ' +
                                  LTBsettings.defaulttech2layoutparams() +
                                  ')'))
    parser_t2l.add_argument('-x', '--excelfile', default=None,
                            help=('the path to the excel file (default: ' +
                                  LTBsettings.tech2layoutparamsexcelfile() +
                                  ')'))
    parser_t2l.add_argument('--nobackup', dest='backup', default=True,
                            action='store_false',
                            help=('Avoids creation of backup files of ' +
                                  'previous output files.'))

    parser_gds = subparsers.add_parser(
            'gdssheet',
            help='Creates the gds translation table file for all technologies')
    parser_gds.add_argument('-o', '--outfile', default=None,
                            help=('the path to the output file (default: ' +
                                  LTBsettings.defaulttech2layoutparams() +
                                  ')'))
    parser_gds.add_argument('-x', '--excelfile', default=None,
                            help=('the path to the excel file (default: ' +
                                  LTBsettings.tech2layoutparamsexcelfile() +
                                  ')'))
    parser_gds.add_argument('--nobackup', dest='backup', default=True,
                            action='store_false',
                            help=('Avoids creation of backup files of ' +
                                  'previous output files.'))

    parser_tdo = subparsers.add_parser('deltdo', help='removes all tdo files')
    parser_tdo.add_argument('-p', '--projectfolder', default='scib',
                            help=('Removes all *.tdo files from ' +
                                  'S:\\Projects\\PROJECTFOLDER\\ and ' +
                                  'subfolders (default: scib)'))
    parser_tdo.add_argument('-d', '--dryrun', default=False,
                            action='store_true',
                            help=('Runs the script but does not really ' +
                                  'remove a file'))
    vias.argparse_setup(subparsers)


def argparse_eval(args):
    # make a dict of args
    dictargs = vars(args)
    funcdict = {'v2common': (v2common,
                             [dictargs.get('technology'),
                              dictargs.get('excelfile'),
                              dictargs.get('backup')]),
                'tech2layoutparams': (tech2layoutparams,
                                      [dictargs.get('excelfile'),
                                       dictargs.get('outfile'),
                                       dictargs.get('backup')]),
                'gdssheet': (gdssheet,
                             [dictargs.get('excelfile'),
                              dictargs.get('outfile'),
                              dictargs.get('backup')]),
                'deltdo': (deltdo,
                           [dictargs.get('projectfolder'),
                            dictargs.get('dryrun')])
                }
    funcdict.update(vias.argparse_eval(args))

    return funcdict


if __name__ == "__main__":
    general.logsetup()
    general.myargparse(argparse_setup, argparse_eval, 'v20240909')
